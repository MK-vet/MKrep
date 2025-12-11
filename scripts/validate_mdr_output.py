#!/usr/bin/env python3
"""
Validation script to verify MDR analysis results.
This script checks that all expected outputs were generated and contain valid data.
"""

import os
import sys
import glob

def check_file_exists(pattern, description):
    """Check if files matching pattern exist."""
    files = glob.glob(pattern)
    if files:
        print(f"✓ {description}: {len(files)} file(s) found")
        for f in files:
            size = os.path.getsize(f) / 1024  # KB
            print(f"  - {os.path.basename(f)} ({size:.1f} KB)")
        return True
    else:
        print(f"✗ {description}: No files found")
        return False

def validate_output():
    """Validate all expected output files."""
    print("=" * 70)
    print("MDR ANALYSIS OUTPUT VALIDATION")
    print("=" * 70)
    
    all_good = True
    
    # Check merged input data
    print("\n1. INPUT DATA:")
    all_good &= check_file_exists("merged_resistance_data.csv", "Merged dataset")
    
    # Check output directory
    print("\n2. OUTPUT DIRECTORY:")
    if os.path.exists("output"):
        print("✓ Output directory exists")
    else:
        print("✗ Output directory not found")
        all_good = False
    
    # Check HTML reports
    print("\n3. HTML REPORTS:")
    all_good &= check_file_exists("output/mdr_analysis_hybrid_*.html", "HTML reports")
    
    # Check Excel reports
    print("\n4. EXCEL REPORTS:")
    all_good &= check_file_exists("output/MDR_Analysis_Hybrid_Report_*.xlsx", "Excel reports")
    
    # Check PNG charts
    print("\n5. PNG VISUALIZATIONS:")
    all_good &= check_file_exists("output/png_charts/*.png", "PNG charts")
    
    # Additional validation with openpyxl if available
    print("\n6. EXCEL CONTENT VALIDATION:")
    try:
        import openpyxl
        excel_files = glob.glob("output/MDR_Analysis_Hybrid_Report_*.xlsx")
        if excel_files:
            wb = openpyxl.load_workbook(excel_files[0], data_only=True)
            expected_sheets = [
                'Metadata', 'Methodology', 'Dataset_Overview', 
                'Freq_Pheno_All', 'Freq_Pheno_MDR',
                'Freq_Gene_All', 'Freq_Gene_MDR',
                'Patterns_Pheno_MDR', 'Patterns_Gene_MDR',
                'Coocc_Pheno_MDR', 'Coocc_Gene_MDR',
                'Gene_Pheno_Assoc', 'Assoc_Rules_Pheno', 'Assoc_Rules_Genes',
                'Network_Edges', 'Network_Communities', 'Network_Stats'
            ]
            actual_sheets = wb.sheetnames
            missing_sheets = [s for s in expected_sheets if s not in actual_sheets]
            
            if not missing_sheets:
                print(f"✓ All {len(expected_sheets)} expected sheets present")
            else:
                print(f"✗ Missing sheets: {', '.join(missing_sheets)}")
                all_good = False
            
            # Check Dataset_Overview has data
            ws = wb['Dataset_Overview']
            if ws.max_row > 5:
                print(f"✓ Dataset_Overview contains data ({ws.max_row} rows)")
            else:
                print("✗ Dataset_Overview appears empty")
                all_good = False
            
            wb.close()
    except ImportError:
        print("⚠ openpyxl not available, skipping detailed Excel validation")
    except Exception as e:
        print(f"⚠ Excel validation error: {e}")
    
    # Summary
    print("\n" + "=" * 70)
    if all_good:
        print("✓ VALIDATION PASSED - All expected outputs present")
        print("\nNext steps:")
        print("1. Open HTML report: output/mdr_analysis_hybrid_*.html")
        print("2. Review Excel report: output/MDR_Analysis_Hybrid_Report_*.xlsx")
        print("3. Check network visualization: output/png_charts/hybrid_network_visualization.png")
    else:
        print("✗ VALIDATION FAILED - Some outputs missing or incomplete")
        print("\nPlease run: python run_mdr_analysis.py merged_resistance_data.csv")
    print("=" * 70)
    
    return all_good

if __name__ == "__main__":
    success = validate_output()
    sys.exit(0 if success else 1)
